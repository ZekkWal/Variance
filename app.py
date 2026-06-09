import os
import streamlit as st
import pandas as pd
import anthropic
import altair as alt
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="FP&A Variance Commentary", layout="wide")
st.title("FP&A Variance Commentary Generator")
st.caption("Upload your budget and actuals CSVs to generate CFO-ready commentary.")

# ── Sample data ───────────────────────────────────────────────────────────────
SAMPLE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")

# (file name, friendly label) for each of the four bundled sample files.
SAMPLE_FILES = {
    "budget":        ("sample_budget.csv",        "Budget — current period"),
    "actuals":       ("sample_actuals.csv",       "Actuals — current period"),
    "prior_budget":  ("sample_prior_budget.csv",  "Budget — prior period"),
    "prior_actuals": ("sample_prior_actuals.csv", "Actuals — prior period"),
}

@st.cache_data
def load_sample_text(file_name):
    """Read a bundled sample CSV as raw text (cached so we only hit disk once)."""
    with open(os.path.join(SAMPLE_DIR, file_name), encoding="utf-8") as f:
        return f.read()

# ── API key input ─────────────────────────────────────────────────────────────
api_key = st.text_input("Enter your Anthropic API Key", type="password")

# ── File upload ───────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    budget_file = st.file_uploader("Upload Budget CSV", type="csv")
with col2:
    actuals_file = st.file_uploader("Upload Actuals CSV", type="csv")

# ── Optional prior period for comparison ──────────────────────────────────────
st.write("**Optional:** Upload prior period data for comparison")
col3, col4 = st.columns(2)
with col3:
    prior_budget_file = st.file_uploader("Upload Prior Period Budget CSV (optional)", type="csv", key="prior_budget")
with col4:
    prior_actuals_file = st.file_uploader("Upload Prior Period Actuals CSV (optional)", type="csv", key="prior_actuals")

# ── Sample data: view / download / load ───────────────────────────────────────
st.markdown("---")
with st.expander("🧪 No data handy? Try the tool with a built-in sample dataset", expanded=False):
    st.markdown(
        "These four files are designed to showcase **every** feature of the tool — "
        "material favorable *and* unfavorable variances, an escalating monthly cost "
        "trend (cloud infrastructure & advertising), and a prior period so the "
        "**period-over-period comparison** lights up. Preview or download any file to "
        "use as a template, or load the whole set into the tool with one click."
    )

    sample_tabs = st.tabs([label for _, label in SAMPLE_FILES.values()])
    for tab, (file_name, label) in zip(sample_tabs, SAMPLE_FILES.values()):
        with tab:
            csv_text = load_sample_text(file_name)
            st.dataframe(pd.read_csv(StringIO(csv_text)), use_container_width=True)
            st.download_button(
                f"⬇️ Download {file_name}",
                data=csv_text,
                file_name=file_name,
                mime="text/csv",
                key=f"download_{file_name}",
            )

    load_col, clear_col = st.columns(2)
    with load_col:
        if st.button("📥 Load sample data into the tool"):
            st.session_state.use_sample = True
            st.rerun()
    with clear_col:
        if st.session_state.get("use_sample") and st.button("✖️ Clear sample data"):
            st.session_state.use_sample = False
            st.rerun()

if st.session_state.get("use_sample"):
    st.success(
        "✅ **Sample dataset loaded** (current *and* prior period). "
        "Scroll down and click **Generate Commentary**. "
        "Any file you upload above will override the matching sample."
    )

# ── Settings ──────────────────────────────────────────────────────────────────
materiality_pct = st.slider(
    "Materiality threshold (only flag variances above this %)",
    min_value=1, max_value=20, value=5
)
audience = st.selectbox("Commentary audience", ["CFO", "Board", "Operations Team"])

# ── Core logic ────────────────────────────────────────────────────────────────
def calculate_variances(budget_df, actuals_df, materiality_pct):
    """
    Do all the math in Python — never let the LLM touch raw numbers.
    Returns a list of flagged variance dicts above the materiality threshold.
    """
    # Sum across all months for each Department + Category row
    month_cols = [
        c for c in budget_df.columns
        if c not in ("Department", "Category", "Total_Budget", "Total_Actuals")
    ]

    budget_df["Total_Budget"] = budget_df[month_cols].sum(axis=1)
    actuals_df["Total_Actuals"] = actuals_df[month_cols].sum(axis=1)

    merged = budget_df[["Department", "Category", "Total_Budget"]].merge(
        actuals_df[["Department", "Category", "Total_Actuals"]],
        on=["Department", "Category"]
    )

    merged["Variance_$"] = merged["Total_Actuals"] - merged["Total_Budget"]
    merged["Variance_%"] = (merged["Variance_$"] / merged["Total_Budget"] * 100).round(1)
    merged["Flag"] = merged["Variance_%"].abs() >= materiality_pct
    merged["Direction"] = merged["Variance_$"].apply(
        lambda x: "Unfavorable" if x > 0 else "Favorable"
    )

    flagged = merged[merged["Flag"]].sort_values("Variance_%", key=abs, ascending=False)
    return merged, flagged

def calculate_monthly_variances(budget_df, actuals_df, flagged_df):
    """
    Calculate monthly variances for flagged line items to show trends.
    Returns a dataframe with monthly breakdown.
    """
    month_cols = [
        c for c in budget_df.columns
        if c not in ("Department", "Category", "Total_Budget", "Total_Actuals")
    ]

    monthly_data = []
    for _, row in flagged_df.iterrows():
        dept = row["Department"]
        cat = row["Category"]
        
        budget_row = budget_df[(budget_df["Department"] == dept) & (budget_df["Category"] == cat)].iloc[0]
        actuals_row = actuals_df[(actuals_df["Department"] == dept) & (actuals_df["Category"] == cat)].iloc[0]
        
        for month in month_cols:
            budget_val = budget_row[month]
            actuals_val = actuals_row[month]
            variance = actuals_val - budget_val
            variance_pct = (variance / budget_val * 100) if budget_val != 0 else 0
            
            monthly_data.append({
                "Department": dept,
                "Category": cat,
                "Month": month,
                "Budget": budget_val,
                "Actuals": actuals_val,
                "Variance_$": variance,
                "Variance_%": variance_pct
            })
    
    return pd.DataFrame(monthly_data)

def compare_periods(current_full_df, prior_full_df):
    """
    Compare variances between two periods.
    Returns dataframe showing current vs prior variance trends.
    """
    current_df = current_full_df[["Department", "Category", "Variance_$", "Variance_%"]].copy()
    current_df.columns = ["Department", "Category", "Current_Variance_$", "Current_Variance_%"]
    
    prior_df = prior_full_df[["Department", "Category", "Variance_$", "Variance_%"]].copy()
    prior_df.columns = ["Department", "Category", "Prior_Variance_$", "Prior_Variance_%"]
    
    comparison = current_df.merge(prior_df, on=["Department", "Category"], how="outer")
    comparison["Variance_Trend_$"] = comparison["Current_Variance_$"] - comparison["Prior_Variance_$"]
    comparison["Variance_Trend_%"] = comparison["Current_Variance_%"] - comparison["Prior_Variance_%"]
    comparison["Trend"] = comparison["Variance_Trend_$"].apply(
        lambda x: "📈 Worsening" if x > 5 else ("📉 Improving" if x < -5 else "➡️ Stable")
    )
    
    return comparison

def validate_data_quality(budget_df, actuals_df):
    """
    Check for data quality issues. Returns a list of warnings/errors.
    """
    issues = []
    
    # Check required columns
    required_cols = ["Department", "Category"]
    for col in required_cols:
        if col not in budget_df.columns:
            issues.append(f"❌ Budget CSV missing '{col}' column")
        if col not in actuals_df.columns:
            issues.append(f"❌ Actuals CSV missing '{col}' column")
    
    if issues:
        return issues
    
    # Check for missing values
    if budget_df[required_cols].isna().any().any():
        issues.append("⚠️ Budget CSV has missing Department or Category values")
    if actuals_df[required_cols].isna().any().any():
        issues.append("⚠️ Actuals CSV has missing Department or Category values")
    
    # Check for duplicate rows
    budget_dups = budget_df[required_cols].duplicated().sum()
    actuals_dups = actuals_df[required_cols].duplicated().sum()
    if budget_dups > 0:
        issues.append(f"⚠️ Budget CSV has {budget_dups} duplicate Department/Category combinations")
    if actuals_dups > 0:
        issues.append(f"⚠️ Actuals CSV has {actuals_dups} duplicate Department/Category combinations")
    
    # Check month columns match
    budget_months = set(c for c in budget_df.columns if c not in required_cols)
    actuals_months = set(c for c in actuals_df.columns if c not in required_cols)
    if budget_months != actuals_months:
        missing_in_actuals = budget_months - actuals_months
        missing_in_budget = actuals_months - budget_months
        if missing_in_actuals:
            issues.append(f"⚠️ Months in Budget but not Actuals: {', '.join(sorted(missing_in_actuals))}")
        if missing_in_budget:
            issues.append(f"⚠️ Months in Actuals but not Budget: {', '.join(sorted(missing_in_budget))}")
    
    # Check for negative budgets
    month_cols = [c for c in budget_df.columns if c not in required_cols]
    if month_cols:
        if (budget_df[month_cols] < 0).any().any():
            issues.append("⚠️ Budget CSV contains negative values")
    
    # Check for mismatched Department/Category between files
    budget_keys = set(zip(budget_df["Department"], budget_df["Category"]))
    actuals_keys = set(zip(actuals_df["Department"], actuals_df["Category"]))
    missing_in_actuals = budget_keys - actuals_keys
    missing_in_budget = actuals_keys - budget_keys
    if missing_in_actuals:
        issues.append(f"⚠️ {len(missing_in_actuals)} Department/Category combinations in Budget but not Actuals")
    if missing_in_budget:
        issues.append(f"⚠️ {len(missing_in_budget)} Department/Category combinations in Actuals but not Budget")
    
    return issues

def export_to_excel(full_df, flagged_df, commentary, audience):
    """
    Generate a polished Excel file with variance summary, flagged items, and commentary.
    """
    output = BytesIO()
    wb = Workbook()
    
    # Sheet 1: Executive Summary
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_exec.column_dimensions["A"].width = 25
    ws_exec.column_dimensions["B"].width = 50
    
    ws_exec["A1"] = "FP&A Variance Commentary"
    ws_exec["A1"].font = Font(size=16, bold=True)
    ws_exec["A3"] = "Audience:"
    ws_exec["B3"] = audience
    ws_exec["A4"] = "Materiality Threshold:"
    ws_exec["B4"] = "See Variance Summary sheet"
    
    ws_exec["A6"] = "Commentary:"
    ws_exec["A6"].font = Font(bold=True, size=12)
    commentary_cell = ws_exec["A7"]
    commentary_cell.value = commentary
    commentary_cell.alignment = Alignment(wrap_text=True)
    ws_exec.row_dimensions[7].height = 200
    
    # Sheet 2: Variance Summary (full)
    ws_summary = wb.create_sheet("Variance Summary")
    display_df = full_df[["Department", "Category", "Total_Budget", "Total_Actuals", "Variance_$", "Variance_%", "Direction", "Flag"]].copy()
    display_df.columns = ["Dept", "Category", "Budget", "Actuals", "Variance $", "Variance %", "Direction", "Flagged"]
    
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            if c_idx >= 3:  # Format as currency/percentage
                if "Variance %" in display_df.columns[c_idx-1]:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '$#,##0'
    
    for col in ws_summary.columns:
        ws_summary.column_dimensions[col[0].column_letter].width = 15
    
    # Sheet 3: Flagged Items (detailed)
    if not flagged_df.empty:
        ws_flagged = wb.create_sheet("Flagged Items")
        flagged_display = flagged_df[["Department", "Category", "Total_Budget", "Total_Actuals", "Variance_$", "Variance_%", "Direction"]].copy()
        flagged_display.columns = ["Dept", "Category", "Budget", "Actuals", "Variance $", "Variance %", "Direction"]
        
        for r_idx, row in enumerate(dataframe_to_rows(flagged_display, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_flagged.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                if r_idx > 1 and c_idx == 7:  # Color code direction
                    if value == "Unfavorable":
                        cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
        
        for col in ws_flagged.columns:
            ws_flagged.column_dimensions[col[0].column_letter].width = 15
    
    wb.save(output)
    output.seek(0)
    return output

def build_prompt(flagged_df, audience):
    """
    Structured prompt — pass verified numbers, ask for language only.
    """
    lines = []
    for _, row in flagged_df.iterrows():
        lines.append(
            f"- {row['Department']} / {row['Category']}: "
            f"Budget ${row['Total_Budget']:,.0f}, "
            f"Actuals ${row['Total_Actuals']:,.0f}, "
            f"Variance ${row['Variance_$']:,.0f} ({row['Variance_%']}%) — {row['Direction']}"
        )

    variance_block = "\n".join(lines)

    prompt = f"""You are a senior FP&A analyst writing a variance commentary for the {audience}.

The following variances have been calculated by our financial system and are mathematically verified.
Do NOT recalculate or second-guess these numbers. Your job is to write clear, professional commentary.

FLAGGED VARIANCES (above materiality threshold):
{variance_block}

Instructions:
1. Write an executive summary (2-3 sentences) of overall budget performance.
2. For each flagged line item, write one concise sentence explaining the variance with a likely business driver.
3. Close with one sentence on recommended management actions.
4. Tone: professional, direct, suitable for {audience}.
5. Do not use bullet points — write in prose paragraphs.
"""
    return prompt

# ── Run button ────────────────────────────────────────────────────────────────
def resolve_source(uploaded_file, sample_key, use_sample):
    """Return a DataFrame from an uploaded file, or the bundled sample if loaded."""
    if uploaded_file is not None:
        return pd.read_csv(uploaded_file)
    if use_sample:
        return pd.read_csv(StringIO(load_sample_text(SAMPLE_FILES[sample_key][0])))
    return None

if st.button("Generate Commentary", type="primary"):
    use_sample = st.session_state.get("use_sample", False)
    budget_df = resolve_source(budget_file, "budget", use_sample)
    actuals_df = resolve_source(actuals_file, "actuals", use_sample)

    if not api_key:
        st.error("Please enter your Anthropic API key.")
    elif budget_df is None or actuals_df is None:
        st.error("Please upload both CSV files, or load the sample dataset above.")
    else:
        with st.spinner("Calculating variances..."):
            
            # Run data quality checks
            quality_issues = validate_data_quality(budget_df, actuals_df)
            if quality_issues:
                with st.expander("⚠️ Data Quality Alerts", expanded=len(quality_issues) > 0):
                    for issue in quality_issues:
                        st.write(issue)
            
            full_df, flagged_df = calculate_variances(budget_df, actuals_df, materiality_pct)

        # Show variance table
        st.subheader("Variance Summary")
        display = full_df[["Department","Category","Total_Budget","Total_Actuals","Variance_$","Variance_%","Direction","Flag"]].copy()
        display.columns = ["Dept","Category","Budget","Actuals","Variance $","Variance %","Direction","Flagged"]

        def highlight_flag(row):
            if row["Flagged"] and row["Direction"] == "Unfavorable":
                return ["background-color: #ffd6d6"] * len(row)
            elif row["Flagged"] and row["Direction"] == "Favorable":
                return ["background-color: #d6f5d6"] * len(row)
            return [""] * len(row)

        st.dataframe(
            display.style.apply(highlight_flag, axis=1).format({
                "Budget": "${:,.0f}",
                "Actuals": "${:,.0f}",
                "Variance $": "${:,.0f}",
                "Variance %": "{:.1f}%"
            }),
            use_container_width=True
        )

        if flagged_df.empty:
            st.info("No variances exceeded the materiality threshold.")
        else:
            # Monthly trends chart
            st.subheader("Monthly Variance Trends")
            monthly_df = calculate_monthly_variances(budget_df, actuals_df, flagged_df)
            monthly_df["Line_Item"] = monthly_df["Department"] + " / " + monthly_df["Category"]
            
            # Chart: monthly variance % by line item
            trend_chart = alt.Chart(monthly_df).mark_line(point=True).encode(
                x=alt.X("Month:N", title="Month"),
                y=alt.Y("Variance_%:Q", title="Variance %"),
                color=alt.Color("Line_Item:N", title="Line Item"),
                tooltip=["Month", "Line_Item", "Variance_%", "Variance_$"]
            ).properties(height=400).interactive()
            
            st.altair_chart(trend_chart, use_container_width=True)
            
            # Variance drivers by department
            st.subheader("Variance Drivers")
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Top Unfavorable Variances**")
                unfavorable = flagged_df[flagged_df["Direction"] == "Unfavorable"].copy()
                if not unfavorable.empty:
                    unfavorable["Label"] = unfavorable["Department"] + " / " + unfavorable["Category"]
                    unfav_chart = alt.Chart(unfavorable.head(10)).mark_bar().encode(
                        y=alt.Y("Label:N", sort="-x"),
                        x=alt.X("Variance_%:Q", title="Variance %"),
                        color=alt.value("#d73027")
                    ).properties(height=300)
                    st.altair_chart(unfav_chart, use_container_width=True)
                else:
                    st.info("No unfavorable variances.")
            
            with col2:
                st.write("**Top Favorable Variances**")
                favorable = flagged_df[flagged_df["Direction"] == "Favorable"].copy()
                if not favorable.empty:
                    favorable["Label"] = favorable["Department"] + " / " + favorable["Category"]
                    fav_chart = alt.Chart(favorable.head(10)).mark_bar().encode(
                        y=alt.Y("Label:N", sort="-x"),
                        x=alt.X("Variance_%:Q", title="Variance %"),
                        color=alt.value("#1a9850")
                    ).properties(height=300)
                    st.altair_chart(fav_chart, use_container_width=True)
                else:
                    st.info("No favorable variances.")
            
            # Period-over-period comparison (if prior data provided)
            prior_budget_df = resolve_source(prior_budget_file, "prior_budget", use_sample)
            prior_actuals_df = resolve_source(prior_actuals_file, "prior_actuals", use_sample)
            if prior_budget_df is not None and prior_actuals_df is not None:
                with st.spinner("Loading prior period data..."):
                    prior_full_df, _ = calculate_variances(prior_budget_df, prior_actuals_df, materiality_pct)
                    
                    comparison_df = compare_periods(full_df, prior_full_df)
                    comparison_df = comparison_df.dropna(subset=["Current_Variance_%", "Prior_Variance_%"])
                    
                    if not comparison_df.empty:
                        st.subheader("Period-over-Period Variance Trends")
                        st.write(f"Current vs Prior: **{comparison_df['Trend'].value_counts().to_dict()}**")
                        
                        # Chart: variance comparison
                        comp_chart_data = comparison_df.copy()
                        comp_chart_data["Line_Item"] = comp_chart_data["Department"] + " / " + comp_chart_data["Category"]
                        comp_chart_data = comp_chart_data.melt(
                            id_vars=["Line_Item"],
                            value_vars=["Current_Variance_%", "Prior_Variance_%"],
                            var_name="Period",
                            value_name="Variance %"
                        )
                        comp_chart_data["Period"] = comp_chart_data["Period"].map({"Current_Variance_%": "Current", "Prior_Variance_%": "Prior"})
                        
                        comp_chart = alt.Chart(comp_chart_data.head(20)).mark_bar().encode(
                            x=alt.X("Line_Item:N", title=""),
                            y=alt.Y("Variance %:Q", title="Variance %"),
                            color=alt.Color("Period:N", title="Period"),
                            tooltip=["Line_Item", "Period", "Variance %"]
                        ).properties(height=400)
                        
                        st.altair_chart(comp_chart, use_container_width=True)
            
            with st.spinner("Generating commentary via Claude..."):
                prompt = build_prompt(flagged_df, audience)
                client = anthropic.Anthropic(api_key=api_key)
                message = client.messages.create(
                    model="claude-opus-4-5",
                    max_tokens=1000,
                    messages=[{"role": "user", "content": prompt}]
                )
                commentary = message.content[0].text

            st.subheader("Generated Commentary")
            st.write(commentary)
            st.code(commentary, language=None)
            st.caption("Copy the text above and paste directly into your board pack or email.")
            
            # Show which variances drove the commentary
            st.subheader("Analysis Sources")
            st.write(f"**{len(flagged_df)} flagged variances** were provided to Claude for commentary generation:")
            
            sources_display = flagged_df[["Department", "Category", "Variance_$", "Variance_%", "Direction"]].copy()
            sources_display.columns = ["Department", "Category", "Variance $", "Variance %", "Direction"]
            sources_display = sources_display.sort_values("Variance %", key=abs, ascending=False)
            
            st.dataframe(
                sources_display.style.format({
                    "Variance $": "${:,.0f}",
                    "Variance %": "{:.1f}%"
                }).map(
                    lambda x: "background-color: #ffe6e6" if x == "Unfavorable" else ("background-color: #e6ffe6" if x == "Favorable" else ""),
                    subset=["Direction"]
                ),
                use_container_width=True
            )
            
            st.caption(
                "⚠️ **Transparency Note:** This commentary was generated by Claude AI based on the flagged variances above. "
                "All calculations are deterministic and reviewed. Commentary should be validated by a financial analyst."
            )
            
            # Export to Excel
            excel_file = export_to_excel(full_df, flagged_df, commentary, audience)
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_file,
                file_name="Variance_Commentary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
